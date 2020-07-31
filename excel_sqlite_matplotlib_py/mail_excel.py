#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

import datetime, string
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date

today = date.today()
now = datetime.datetime.now()

# current month name format
monthWord = now.strftime('%B')

# workbook name
wbFilename = (f'{monthWord}_{today.year}.xlsx')

# name of worksheet
sheetWb = (f'{today.month}-{today.day}')

wb = Workbook()

sheetOne = wb.active
sheetOne.title = sheetWb

thin = Side(border_style="thin", color="000000")

# all uppercase letters
upAlpha = list(string.ascii_uppercase)

yCol = []
'''for col in range(0,7):
	yCol = upAlpha[col]
	for row in range(1,71):
		c = sheetOne[f'{yCol}{row}']
		c.value = ''
		c.border = Border(top=thin, right=thin, bottom=thin, left=thin)
		c.alignment = Alignment(horizontal="center", vertical="center")
	'''

row_count = len(sheetOne['A'])
sheetOne.insert_rows(row_count+1)
print(row_count)
wb.save(filename = wbFilename)

# method to print list
'''head = (
	('Date', 'User', 'tracking'),
	('a', 'b', 'c')
	)
for row in head:
	sheetOne.append(row)''' 

