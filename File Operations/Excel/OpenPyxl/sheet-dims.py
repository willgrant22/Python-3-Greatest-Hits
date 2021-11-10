#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import Workbook, load_workbook

workbook = load_workbook(filename="Order Management-CDM - DC Sync.xlsx")
sheet = workbook.active

rows = sheet.max_row
columns = sheet.max_column

print(sheet.dimensions, rows, columns, len(sheet['1']))