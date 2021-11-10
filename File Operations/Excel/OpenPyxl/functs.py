#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
print(wb, ws)

ws.title = "New Title"
ws.sheet_properties.tabColor = "1072BA"
ws3 = wb["New Title"]

print(wb.sheetnames)

for sheet in wb:
	print(sheet.title)

source = wb.active
target = wb.copy_worksheet(source)

d = ws.cell(row=4, column=2, value=10)
c = ws['A4']

cell_range = ws['A1':'C2']

colC = ws['C']
col_range = ws['C:D']
row10 = ws[10]
row_range = ws[5:10]

print(col_range,row_range,cell_range)
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
   for cell in row:
       print(cell)


for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
   print(row)

c.value = 'hello, world'
print(c.value)


d.value = 3.14
print(d.value)

wb = Workbook()
wb.save('balances.xlsx')