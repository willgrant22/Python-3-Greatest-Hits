#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import load_workbook, Workbook

wb = load_workbook(filename='sample.xlsx')
w = Workbook()

wsNames = wb.sheetnames

ws = wb.active

minR = ws.min_row
maxR = ws.max_row

minC = ws.min_column
maxC = ws.max_column

for row in ws.iter_rows(min_row=minR,
                           max_row=maxR,
                           min_col=minC,
                           max_col=maxC,
                           values_only=True):
    print(row)

print(wsNames,w.get_index(ws.title[0]))


