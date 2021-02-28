#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

import json
from openpyxl import load_workbook

workbook = load_workbook(filename="OOM.xlsx")
sheet = workbook.active

products = {}

# Using the values_only because you want to return the cells' values
for row in sheet.iter_rows(min_row=2,
                           min_col=4,
                           max_col=7,
                           values_only=True):
    product_id = row[0]
    product = {
        "parent": row[1],
        "title": row[2],
        "category": row[3]
    }
    products[product_id] = product

# Using json here to be able to format the output for displaying later
print(json.dumps(products))