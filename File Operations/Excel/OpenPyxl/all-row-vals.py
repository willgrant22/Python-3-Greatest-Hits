#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import load_workbook
workbook = load_workbook(filename="Order Management-CDM - DC Sync.xlsx")
print(workbook.sheetnames)

sheet = workbook.active
print(sheet)


print(sheet.title)

for value in sheet.iter_rows(values_only=True):
    print(value)