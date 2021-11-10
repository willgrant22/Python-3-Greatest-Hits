#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import load_workbook
workbook = load_workbook(filename="OOM.xlsx")
print(workbook.sheetnames)

sheet = workbook.active
print(sheet)


print(sheet.title)

print(sheet["A1"])


print(sheet["A1"].value)


print(sheet["F10"].value)
