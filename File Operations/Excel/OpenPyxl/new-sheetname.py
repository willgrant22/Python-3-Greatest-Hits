#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import Workbook, load_workbook

workbook = load_workbook(filename="OOM.xlsx")
sheet = workbook.active

print(workbook.sheetnames)

products_sheet = workbook["OM"]

products_sheet.title = "New OM"

print(workbook.sheetnames)