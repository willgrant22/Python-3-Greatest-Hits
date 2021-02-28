#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import Workbook, load_workbook

workbook = load_workbook(filename="OOM.xlsx")
sheet = workbook.active

sheet.insert_cols(idx=1)

for row in sheet.iter_rows(values_only=True):
  print(row)

# Insert a column before the existing column 1 ("A")
sheet.insert_cols(idx=1)

for row in sheet.iter_rows(values_only=True):
  print(row)

# Insert 5 columns between column 2 ("B") and 3 ("C")
sheet.insert_cols(idx=3, amount=5)

for row in sheet.iter_rows(values_only=True):
  print(row)

# Delete the created columns
sheet.delete_cols(idx=3, amount=5)
sheet.delete_cols(idx=1)

for row in sheet.iter_rows(values_only=True):
  print(row)


# Insert a new row in the beginning
sheet.insert_rows(idx=1)

for row in sheet.iter_rows(values_only=True):
  print(row)

# Insert 3 new rows in the beginning
sheet.insert_rows(idx=1, amount=3)

for row in sheet.iter_rows(values_only=True):
  print(row)

# Delete the first 4 rows
sheet.delete_rows(idx=1, amount=4)

for row in sheet.iter_rows(values_only=True):
  print(row)
#workbook.save(filename=filename)