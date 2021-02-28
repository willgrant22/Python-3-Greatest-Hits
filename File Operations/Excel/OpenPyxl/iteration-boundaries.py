#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import load_workbook
workbook = load_workbook(filename="Order Management-CDM - DC Sync.xlsx")
print(workbook.sheetnames)

sheet = workbook.active
print(sheet)


print(sheet.title)

for row in sheet.iter_rows(min_row=1,
                           max_row=2,
                           min_col=1,
                           max_col=3):
    print(row)

for column in sheet.iter_cols(min_row=1,
                              max_row=2,
                              min_col=1,
                              max_col=3):
    print(column)


print(sheet.dimensions)
print("Minimum row: {0}".format(sheet.min_row))
print("Maximum row: {0}".format(sheet.max_row))
print("Minimum column: {0}".format(sheet.min_column))
print("Maximum column: {0}".format(sheet.max_column))