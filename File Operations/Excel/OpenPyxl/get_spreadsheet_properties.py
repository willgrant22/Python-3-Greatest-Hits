#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color,Font,Alignment,PatternFill,Border,Side,Protection

class Excel:
	def __init__(self, Workbook, Spreadsheet):
		self.Dict = {}
		self.Workbook = Workbook
		self.Spreadsheet = Spreadsheet
		self.WorkbookInstance = load_workbook(filename=self.Workbook)
		self.SpreadsheetInstance = self.WorkbookInstance.active
		self.SpreadsheetDimensions = self.SpreadsheetInstance.dimensions
		self.Dict['Workbook'] = self.Workbook
		self.Dict['Sheet'] = self.Spreadsheet
		self.Dict['Dimensions'] = self.SpreadsheetDimensions

	def GetInstance(self):
		return self.Dict

FirstDocument = Excel('A.xlsx', 'A-Sheet')
FirstDocument.GetInstance()

print(FirstDocument.Dict)

SecondDocument = Excel('B.xlsx', 'B-Sheet')
SecondDocument.GetInstance()

print(SecondDocument.Dict)
