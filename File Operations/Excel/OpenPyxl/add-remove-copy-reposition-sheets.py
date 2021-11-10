#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import Workbook, load_workbook

workbook = load_workbook(filename="OOM.xlsx")
sheet = workbook.active

print(workbook.sheetnames)

products_sheet = workbook["OM"]

products_sheet.title = "New OM"

print(workbook.sheetnames)

operations_sheet = workbook.create_sheet("Operations")


print(workbook.sheetnames)

hr_sheet = workbook.create_sheet("HR", 0)


print(workbook.sheetnames)

workbook.remove(operations_sheet)

print(workbook.sheetnames)

workbook.remove(hr_sheet)

products_sheet = workbook["Sheet1"]
print(workbook.copy_worksheet(products_sheet))

print(workbook.sheetnames)