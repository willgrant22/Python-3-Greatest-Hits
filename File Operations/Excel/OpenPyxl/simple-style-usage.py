#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, Alignment, PatternFill, Border, Side
#workbook = load_workbook(filename="Order Management-CDM - DC Sync.xlsx")
workbook = Workbook()
print(workbook.sheetnames)

sheet = workbook.active

minR = sheet.min_row
maxR = sheet.max_row

minC = sheet.min_column
maxC = sheet.max_column

c = sheet['A1'] = 'Test'
c = sheet['A1']

bd = Side(style='thin', color="000000")
c.border = Border(top=bd, left=bd, right=bd, bottom=bd)

c.font = Font(name='Times New Roman', size=16, color='00000000', bold=True)
c.alignment = Alignment(horizontal='center',vertical='center')

c.fill = PatternFill("solid", fgColor="FFFF00")


workbook.save("om.xlsx")