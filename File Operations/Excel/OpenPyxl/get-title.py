#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import load_workbook
workbook = load_workbook(filename="OOM.xlsx")
print(workbook.sheetnames)

sheet = workbook.active
print(sheet)


print(sheet.title)

for value in sheet.iter_rows(min_row=1,
                             max_row=1,
                             values_only=True):
    print(value)