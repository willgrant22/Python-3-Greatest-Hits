#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#create excel type item
wb = Workbook()
# select the active workws
ws = wb.active

counter = 0
for column in range(1,6):
    column_letter = get_column_letter(column)
    for row in range(1,11):
        counter = counter +1
        ws[column_letter + str(row)] = counter


minR = ws.min_row
maxR = ws.max_row

minC = ws.min_column
maxC = ws.max_column

print(ws)
print(ws.title)
print(ws.dimensions)
print(f"Minimum row: {minR}")
print(f"Maximum row: {maxR}\n")
print(f"Minimum column: {minC}")
print(f"Maximum column: {maxC}")

for column in range(minC, maxC + 1):
    column_letter = get_column_letter(column)
    print(('\n' + column_letter * 2))
    for row in range(minR,maxR + 1):
        counter = counter +1
        ws[column_letter + str(row)] = counter
        print(row)
print(column_letter+str(row))

c = ws[f'{column_letter+str(row)}'] = 'Test'

wb.save("sample.xlsx")