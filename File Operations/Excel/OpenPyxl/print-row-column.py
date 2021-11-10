#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import load_workbook
workbook = load_workbook(filename="OOM.xlsx")
print(workbook.sheetnames)

sheet = workbook.active
print(sheet)


print(sheet.title)

x = sheet.cell(row=10, column=6)


y = sheet.cell(row=10, column=6).value

print(f'{x}, {y}')
