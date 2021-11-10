#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font

book = Workbook()
sheet = book.active

rows = (
    (34, 26),
    (88, 36),
    (24, 29),
    (15, 22),
    (56, 13),
    (76, 18)
)

for row in rows:
    sheet.append(row)

cell = sheet.cell(row=7, column=2)
cell.value = "=SUM(A1:B6)"
cell.font = Font(bold=True)

book.save('formulas.xlsx')