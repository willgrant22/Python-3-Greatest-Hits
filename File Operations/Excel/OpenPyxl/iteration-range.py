#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import load_workbook

workbook = load_workbook(filename="Order Management-CDM - DC Sync.xlsx")

print(workbook.sheetnames)

sheet = workbook.active

minR = sheet.min_row
maxR = sheet.max_row

minC = sheet.min_column
maxC = sheet.max_column

print(sheet)
print(sheet.title)
print(sheet.dimensions)
print(f"Minimum row: {minR}")
print(f"Maximum row: {maxR}\n")
print(f"Minimum column: {minC}")
print(f"Maximum column: {maxC}")

for value in sheet.iter_rows(min_row=minR,
                             max_row=maxR,
                             min_col=minC,
                             max_col=maxC,
                             values_only=True):
    print(value)

x = 0
a = 0
y = 7210434923
z = []
for col in sheet['A']:
	print(col.value)
	a = a + 1
	if col.value == y:
		x = x + 1
		z.append(a)

print(f'\n{y} occurs: {x} times\nrows of occurance: {z[0]}')
