#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# Author :  Will Grant
# =============================================================================

from openpyxl import Workbook, load_workbook

workbook = load_workbook(filename="OOM.xlsx")
sheet = workbook.active

sheet.insert_cols(idx=1)

for row in sheet.iter_rows(values_only=True):
  print(row)

#workbook.save(filename=filename)